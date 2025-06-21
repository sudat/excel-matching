"""
Excel操作のユーティリティ関数
"""

import logging
from typing import List, Dict, Any
from io import BytesIO
import openpyxl
import pandas as pd
from fastapi import HTTPException

logger = logging.getLogger(__name__)


def get_excel_sheets_info(file_content: bytes) -> List[Dict[str, Any]]:
    """Excelファイルからシート情報を取得"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        sheets_info = []

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # シートの最大行・列を取得
            max_row = sheet.max_row
            max_col = sheet.max_column

            # データの有無を確認（最初の100行をサンプリング）
            has_data = False
            data_cells = 0
            sample_rows = min(100, max_row) if max_row else 0
            sample_cols = min(20, max_col) if max_col else 0

            if sample_rows > 0 and sample_cols > 0:
                for row in range(1, sample_rows + 1):
                    for col in range(1, sample_cols + 1):
                        cell_value = sheet.cell(row, col).value
                        if cell_value is not None and str(cell_value).strip() != "":
                            data_cells += 1
                            has_data = True

                # データ密度を計算
                total_sample_cells = sample_rows * sample_cols
                data_density = (
                    data_cells / total_sample_cells if total_sample_cells > 0 else 0
                )
            else:
                data_density = 0

            # データ範囲を推定
            data_range = None
            if has_data and max_row > 0 and max_col > 0:
                end_col_letter = openpyxl.utils.get_column_letter(max_col)
                data_range = f"A1:{end_col_letter}{max_row}"

            sheet_info = {
                "name": sheet_name,
                "row_count": max_row if max_row else 0,
                "col_count": max_col if max_col else 0,
                "has_data": has_data,
                "data_range": data_range,
                "data_density": round(data_density, 3),
                "estimated_data_cells": data_cells,
            }
            sheets_info.append(sheet_info)

        workbook.close()
        return sheets_info

    except Exception as e:
        logger.error(f"Error reading Excel sheets: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Excelファイルのシート情報取得中にエラーが発生しました: {str(e)}",
        )


def extract_table_data(
    workbook_data: bytes, sheet_name: str, table_info: Dict[str, Any]
) -> Dict[str, Any]:
    """指定された表の全データを抽出する"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(workbook_data))
        sheet = workbook[sheet_name]

        range_info = table_info["range"]
        start_row = range_info["start_row"]
        end_row = range_info["end_row"]
        start_col = range_info["start_col"]
        end_col = range_info["end_col"]
        header_row = table_info.get("header_row")
        headers = table_info.get("headers", [])

        # 全データを収集
        all_data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row, col).value
                row_data.append(cell_value)
            all_data.append(row_data)

        # ヘッダー行がある場合はデータ部分のみ抽出
        records = []
        data_start_idx = 1 if header_row else 0

        for row_idx in range(data_start_idx, len(all_data)):
            record = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(all_data[row_idx]):
                    value = all_data[row_idx][col_idx]
                    record[header] = str(value) if value is not None else ""
                else:
                    record[header] = ""
            records.append(record)

        # データ型分析（サービスからインポートを避けるため簡易実装）
        if records:
            df = pd.DataFrame(records)
            # data_types とquality_reportは別のモジュールから取得すべきだが、
            # 循環インポートを避けるため、ここでは簡易的に空のdictを返す
            data_types = {}
            quality_report = {}
        else:
            data_types = {}
            quality_report = {}

        workbook.close()

        return {
            "headers": headers,
            "records": records,
            "data_types": data_types,
            "quality_info": quality_report,
            "total_records": len(records),
        }

    except Exception as e:
        logger.error(f"Error extracting table data: {str(e)}")
        raise
