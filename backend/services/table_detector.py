"""
表検出器の実装
"""

from abc import ABC, abstractmethod
from typing import List, Dict, Any, Optional
import logging
from io import BytesIO
import re
import openpyxl

from models.table_models import TableCandidate

logger = logging.getLogger(__name__)


class TableDetector(ABC):
    """表検出器の抽象基底クラス - 将来的なLLM置き換えに対応"""

    @abstractmethod
    def detect_tables(
        self,
        workbook_data: bytes,
        sheet_name: str,
        min_rows: int = 3,
        min_cols: int = 2,
        max_tables: int = 10,
    ) -> List[TableCandidate]:
        """
        指定されたシート内の表を検出する

        Args:
            workbook_data: Excelワークブックのバイナリデータ
            sheet_name: 検出対象のシート名
            min_rows: 表として認識する最小行数
            min_cols: 表として認識する最小列数
            max_tables: 検出する表の最大数

        Returns:
            検出された表候補のリスト（品質スコア降順）
        """
        pass

    @abstractmethod
    def get_detector_info(self) -> Dict[str, Any]:
        """検出器の情報を返す"""
        pass


class StatisticalTableDetector(TableDetector):
    """統計的手法による表検出器"""

    def __init__(self):
        self.name = "Statistical Table Detector"
        self.version = "1.0.0"

    def get_detector_info(self) -> Dict[str, Any]:
        """検出器の情報を返す"""
        return {
            "name": self.name,
            "version": self.version,
            "type": "statistical",
            "description": "統計的手法により表を検出。データ密度、連続性、ヘッダー特徴を分析",
        }

    def detect_tables(
        self,
        workbook_data: bytes,
        sheet_name: str,
        min_rows: int = 3,
        min_cols: int = 2,
        max_tables: int = 10,
    ) -> List[TableCandidate]:
        """統計的手法で表を検出"""
        try:
            workbook = openpyxl.load_workbook(BytesIO(workbook_data), read_only=True)
            sheet = workbook[sheet_name]

            # データ領域を分析
            data_regions = self._find_data_regions(sheet)
            table_candidates = []

            for region_id, region in enumerate(data_regions):
                # 各領域を表候補として評価
                candidate = self._analyze_data_region(
                    sheet, region, f"table_{region_id + 1}", sheet_name
                )

                if (
                    candidate
                    and candidate.row_count >= min_rows
                    and candidate.col_count >= min_cols
                ):
                    table_candidates.append(candidate)

            # 品質スコアでソート
            table_candidates.sort(key=lambda x: x.quality_score, reverse=True)

            workbook.close()
            return table_candidates[:max_tables]

        except Exception as e:
            logger.error(f"Error detecting tables in sheet {sheet_name}: {str(e)}")
            return []

    def _find_data_regions(self, sheet) -> List[Dict[str, int]]:
        """データ領域を検出する"""
        regions = []
        max_row = sheet.max_row or 1
        max_col = sheet.max_column or 1

        # データ密度マップを作成
        data_map = {}
        for row in range(1, min(max_row + 1, 201)):  # 最大200行まで分析
            for col in range(1, min(max_col + 1, 51)):  # 最大50列まで分析
                cell_value = sheet.cell(row, col).value
                has_data = (
                    cell_value is not None
                    and str(cell_value).strip() != ""
                    and str(cell_value).strip() != "0"
                )
                data_map[(row, col)] = has_data

        # 連続するデータブロックを検出
        visited = set()

        for row in range(1, min(max_row + 1, 201)):
            for col in range(1, min(max_col + 1, 51)):
                if (row, col) not in visited and data_map.get((row, col), False):
                    region = self._expand_data_region(
                        data_map, row, col, visited, max_row, max_col
                    )
                    if region and (region["end_row"] - region["start_row"] + 1) >= 3:
                        regions.append(region)

        return regions

    def _expand_data_region(
        self, data_map, start_row, start_col, visited, max_row, max_col
    ):
        """データ領域を拡張する"""
        # 領域の境界を探索
        min_row, max_row_found = start_row, start_row
        min_col, max_col_found = start_col, start_col

        # 行方向の拡張
        for row in range(start_row, min(max_row + 1, 201)):
            has_data_in_row = False
            for col in range(start_col, min(max_col + 1, 51)):
                if data_map.get((row, col), False):
                    has_data_in_row = True
                    max_col_found = max(max_col_found, col)
                    break
            if has_data_in_row:
                max_row_found = row
            else:
                # 連続する2行以上空行があったら終了
                if row > start_row + 1:
                    break

        # 列方向の拡張
        for col in range(start_col, min(max_col + 1, 51)):
            has_data_in_col = False
            for row in range(start_row, max_row_found + 1):
                if data_map.get((row, col), False):
                    has_data_in_col = True
                    break
            if has_data_in_col:
                max_col_found = col
            else:
                break

        # 訪問済みマークを設定
        for row in range(min_row, max_row_found + 1):
            for col in range(min_col, max_col_found + 1):
                visited.add((row, col))

        return {
            "start_row": min_row,
            "end_row": max_row_found,
            "start_col": min_col,
            "end_col": max_col_found,
        }

    def _analyze_data_region(
        self, sheet, region, table_id, sheet_name
    ) -> Optional[TableCandidate]:
        """データ領域を表として分析する"""
        try:
            start_row = region["start_row"]
            end_row = region["end_row"]
            start_col = region["start_col"]
            end_col = region["end_col"]

            row_count = end_row - start_row + 1
            col_count = end_col - start_col + 1

            # データ収集
            data_matrix = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row, col).value
                    row_data.append(cell_value)
                data_matrix.append(row_data)

            # ヘッダー行を検出
            header_row_idx = self._detect_header_row(data_matrix)
            header_row = (
                start_row + header_row_idx if header_row_idx is not None else None
            )

            # ヘッダー取得
            headers = []
            if header_row_idx is not None:
                header_data = data_matrix[header_row_idx]
                headers = [
                    str(cell) if cell is not None else f"列{i+1}"
                    for i, cell in enumerate(header_data)
                ]
            else:
                headers = [f"列{i+1}" for i in range(col_count)]

            # サンプルデータ取得（ヘッダー + 3行）
            sample_data = []
            data_start_idx = (header_row_idx + 1) if header_row_idx is not None else 0

            for i in range(data_start_idx, min(data_start_idx + 3, len(data_matrix))):
                if i < len(data_matrix):
                    row_dict = {}
                    for j, header in enumerate(headers):
                        if j < len(data_matrix[i]):
                            value = data_matrix[i][j]
                            row_dict[header] = str(value) if value is not None else ""
                        else:
                            row_dict[header] = ""
                    sample_data.append(row_dict)

            # データ密度計算
            total_cells = row_count * col_count
            data_cells = sum(
                1
                for row in data_matrix
                for cell in row
                if cell is not None and str(cell).strip() != ""
            )
            data_density = data_cells / total_cells if total_cells > 0 else 0

            # 品質スコア計算
            quality_score = self._calculate_quality_score(
                data_matrix, row_count, col_count, data_density, header_row_idx
            )

            # 推定レコード数
            estimated_records = (
                row_count - 1 if header_row_idx is not None else row_count
            )

            return TableCandidate(
                table_id=table_id,
                sheet_name=sheet_name,
                start_row=start_row,
                end_row=end_row,
                start_col=start_col,
                end_col=end_col,
                header_row=header_row,
                quality_score=quality_score,
                data_density=data_density,
                row_count=row_count,
                col_count=col_count,
                estimated_records=estimated_records,
                headers=headers,
                sample_data=sample_data,
                metadata={
                    "detection_method": "statistical",
                    "data_cells": data_cells,
                    "total_cells": total_cells,
                },
            )

        except Exception as e:
            logger.error(f"Error analyzing data region: {str(e)}")
            return None

    def _detect_header_row(self, data_matrix) -> Optional[int]:
        """ヘッダー行を検出する"""
        if len(data_matrix) < 2:
            return None

        best_header_idx = None
        best_score = 0

        # 最初の3行までをヘッダー候補として分析
        for row_idx in range(min(3, len(data_matrix))):
            row = data_matrix[row_idx]

            # 文字列の割合を計算
            string_count = 0
            non_empty_count = 0

            for cell in row:
                if cell is not None and str(cell).strip() != "":
                    non_empty_count += 1
                    # 数値でない場合は文字列とみなす
                    try:
                        float(str(cell).replace(",", ""))
                    except (ValueError, TypeError):
                        string_count += 1

            if non_empty_count == 0:
                continue

            string_ratio = string_count / non_empty_count

            # ヘッダースコア（文字列率 + データ充填率）
            fill_ratio = non_empty_count / len(row)
            header_score = string_ratio * 0.7 + fill_ratio * 0.3

            if header_score > best_score and string_ratio >= 0.5:
                best_score = header_score
                best_header_idx = row_idx

        return best_header_idx

    def _calculate_quality_score(
        self, data_matrix, row_count, col_count, data_density, header_row_idx
    ) -> float:
        """表の品質スコアを計算"""
        score = 0.0

        # データ密度による評価 (0-0.3)
        density_score = min(data_density * 3, 0.3)
        score += density_score

        # サイズによる評価 (0-0.3)
        size_score = min((row_count * col_count) / 100, 0.3)
        score += size_score

        # ヘッダーの有無による評価 (0-0.2)
        if header_row_idx is not None:
            score += 0.2

        # データの一貫性による評価 (0-0.2)
        consistency_score = self._calculate_data_consistency(
            data_matrix, header_row_idx
        )
        score += consistency_score

        return min(score, 1.0)

    def _calculate_data_consistency(self, data_matrix, header_row_idx) -> float:
        """データの一貫性スコアを計算"""
        if len(data_matrix) <= 1:
            return 0.0

        data_start = (header_row_idx + 1) if header_row_idx is not None else 0
        if data_start >= len(data_matrix):
            return 0.0

        col_consistency_scores = []

        for col_idx in range(len(data_matrix[0])):
            col_data = []
            for row_idx in range(data_start, len(data_matrix)):
                if col_idx < len(data_matrix[row_idx]):
                    cell = data_matrix[row_idx][col_idx]
                    if cell is not None and str(cell).strip() != "":
                        col_data.append(cell)

            if len(col_data) < 2:
                continue

            # データ型の一貫性を確認
            numeric_count = 0
            date_count = 0

            for value in col_data:
                str_value = str(value).strip()

                # 数値チェック
                try:
                    float(str_value.replace(",", ""))
                    numeric_count += 1
                    continue
                except (ValueError, TypeError):
                    pass

                # 日付チェック
                if re.match(r"\d{4}[-/]\d{1,2}[-/]\d{1,2}", str_value):
                    date_count += 1

            total_values = len(col_data)
            if total_values > 0:
                numeric_ratio = numeric_count / total_values
                date_ratio = date_count / total_values

                # 一貫性スコア（1種類のデータ型が80%以上なら高評価）
                if numeric_ratio >= 0.8 or date_ratio >= 0.8:
                    col_consistency_scores.append(1.0)
                elif numeric_ratio >= 0.6 or date_ratio >= 0.6:
                    col_consistency_scores.append(0.6)
                else:
                    col_consistency_scores.append(0.3)

        if not col_consistency_scores:
            return 0.0

        avg_consistency = sum(col_consistency_scores) / len(col_consistency_scores)
        return min(avg_consistency * 0.2, 0.2)
