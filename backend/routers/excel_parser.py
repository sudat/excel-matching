from fastapi import APIRouter, UploadFile, File, HTTPException, Path
from typing import Dict, Any, Optional, List, Protocol, Tuple
import pandas as pd
import openpyxl
import uuid
import os
import logging
from io import BytesIO
from datetime import datetime, timedelta
import re
import numpy as np
from collections import Counter
from abc import ABC, abstractmethod

# ログ設定
logger = logging.getLogger(__name__)
logger.setLevel(logging.INFO)

router = APIRouter(prefix="/api", tags=["Excel Parser"])

# 表検出関連のデータクラス
class TableCandidate:
    """検出された表の候補を表すクラス"""
    def __init__(
        self,
        table_id: str,
        sheet_name: str,
        start_row: int,
        end_row: int,
        start_col: int,
        end_col: int,
        header_row: Optional[int] = None,
        quality_score: float = 0.0,
        data_density: float = 0.0,
        row_count: int = 0,
        col_count: int = 0,
        estimated_records: int = 0,
        headers: Optional[List[str]] = None,
        sample_data: Optional[List[Dict[str, Any]]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ):
        self.table_id = table_id
        self.sheet_name = sheet_name
        self.start_row = start_row
        self.end_row = end_row
        self.start_col = start_col
        self.end_col = end_col
        self.header_row = header_row
        self.quality_score = quality_score
        self.data_density = data_density
        self.row_count = row_count
        self.col_count = col_count
        self.estimated_records = estimated_records
        self.headers = headers or []
        self.sample_data = sample_data or []
        self.metadata = metadata or {}
    
    def to_dict(self) -> Dict[str, Any]:
        """辞書形式に変換"""
        return {
            'table_id': self.table_id,
            'sheet_name': self.sheet_name,
            'range': {
                'start_row': self.start_row,
                'end_row': self.end_row,
                'start_col': self.start_col,
                'end_col': self.end_col,
                'excel_range': f"{openpyxl.utils.get_column_letter(self.start_col)}{self.start_row}:{openpyxl.utils.get_column_letter(self.end_col)}{self.end_row}"
            },
            'header_row': self.header_row,
            'quality_score': round(self.quality_score, 3),
            'data_density': round(self.data_density, 3),
            'dimensions': {
                'row_count': self.row_count,
                'col_count': self.col_count,
                'estimated_records': self.estimated_records
            },
            'headers': self.headers,
            'sample_data': self.sample_data,
            'metadata': self.metadata
        }

class TableDetector(ABC):
    """表検出器の抽象基底クラス - 将来的なLLM置き換えに対応"""
    
    @abstractmethod
    def detect_tables(
        self, 
        workbook_data: bytes, 
        sheet_name: str,
        min_rows: int = 3,
        min_cols: int = 2,
        max_tables: int = 10
    ) -> List[TableCandidate]:
        """
        指定されたシート内の表を検出する
        
        Args:
            workbook_data: Excelワークブックのバイナリデータ
            sheet_name: 検出対象のシート名
            min_rows: 表として認識する最小行数
            min_cols: 表として認識する最小列数
            max_tables: 検出する表の最大数
            
        Returns:
            検出された表候補のリスト（品質スコア降順）
        """
        pass
    
    @abstractmethod
    def get_detector_info(self) -> Dict[str, Any]:
        """検出器の情報を返す"""
        pass

class StatisticalTableDetector(TableDetector):
    """統計的手法による表検出器"""
    
    def __init__(self):
        self.name = "Statistical Table Detector"
        self.version = "1.0.0"
    
    def get_detector_info(self) -> Dict[str, Any]:
        """検出器の情報を返す"""
        return {
            "name": self.name,
            "version": self.version,
            "type": "statistical",
            "description": "統計的手法により表を検出。データ密度、連続性、ヘッダー特徴を分析"
        }
    
    def detect_tables(
        self, 
        workbook_data: bytes, 
        sheet_name: str,
        min_rows: int = 3,
        min_cols: int = 2,
        max_tables: int = 10
    ) -> List[TableCandidate]:
        """統計的手法で表を検出"""
        try:
            workbook = openpyxl.load_workbook(BytesIO(workbook_data), read_only=True)
            sheet = workbook[sheet_name]
            
            # データ領域を分析
            data_regions = self._find_data_regions(sheet)
            table_candidates = []
            
            for region_id, region in enumerate(data_regions):
                # 各領域を表候補として評価
                candidate = self._analyze_data_region(
                    sheet, region, f"table_{region_id + 1}", sheet_name
                )
                
                if candidate and candidate.row_count >= min_rows and candidate.col_count >= min_cols:
                    table_candidates.append(candidate)
            
            # 品質スコアでソート
            table_candidates.sort(key=lambda x: x.quality_score, reverse=True)
            
            workbook.close()
            return table_candidates[:max_tables]
            
        except Exception as e:
            logger.error(f"Error detecting tables in sheet {sheet_name}: {str(e)}")
            return []
    
    def _find_data_regions(self, sheet) -> List[Dict[str, int]]:
        """データ領域を検出する"""
        regions = []
        max_row = sheet.max_row or 1
        max_col = sheet.max_column or 1
        
        # データ密度マップを作成
        data_map = {}
        for row in range(1, min(max_row + 1, 201)):  # 最大200行まで分析
            for col in range(1, min(max_col + 1, 51)):  # 最大50列まで分析
                cell_value = sheet.cell(row, col).value
                has_data = (cell_value is not None and 
                           str(cell_value).strip() != '' and 
                           str(cell_value).strip() != '0')
                data_map[(row, col)] = has_data
        
        # 連続するデータブロックを検出
        visited = set()
        
        for row in range(1, min(max_row + 1, 201)):
            for col in range(1, min(max_col + 1, 51)):
                if (row, col) not in visited and data_map.get((row, col), False):
                    region = self._expand_data_region(data_map, row, col, visited, max_row, max_col)
                    if region and (region['end_row'] - region['start_row'] + 1) >= 3:
                        regions.append(region)
        
        return regions
    
    def _expand_data_region(self, data_map, start_row, start_col, visited, max_row, max_col):
        """データ領域を拡張する"""
        # 領域の境界を探索
        min_row, max_row_found = start_row, start_row
        min_col, max_col_found = start_col, start_col
        
        # 行方向の拡張
        for row in range(start_row, min(max_row + 1, 201)):
            has_data_in_row = False
            for col in range(start_col, min(max_col + 1, 51)):
                if data_map.get((row, col), False):
                    has_data_in_row = True
                    max_col_found = max(max_col_found, col)
                    break
            if has_data_in_row:
                max_row_found = row
            else:
                # 連続する2行以上空行があったら終了
                if row > start_row + 1:
                    break
        
        # 列方向の拡張
        for col in range(start_col, min(max_col + 1, 51)):
            has_data_in_col = False
            for row in range(start_row, max_row_found + 1):
                if data_map.get((row, col), False):
                    has_data_in_col = True
                    break
            if has_data_in_col:
                max_col_found = col
            else:
                break
        
        # 訪問済みマークを設定
        for row in range(min_row, max_row_found + 1):
            for col in range(min_col, max_col_found + 1):
                visited.add((row, col))
        
        return {
            'start_row': min_row,
            'end_row': max_row_found,
            'start_col': min_col,
            'end_col': max_col_found
        }
    
    def _analyze_data_region(self, sheet, region, table_id, sheet_name) -> Optional[TableCandidate]:
        """データ領域を表として分析する"""
        try:
            start_row = region['start_row']
            end_row = region['end_row']
            start_col = region['start_col']
            end_col = region['end_col']
            
            row_count = end_row - start_row + 1
            col_count = end_col - start_col + 1
            
            # データ収集
            data_matrix = []
            for row in range(start_row, end_row + 1):
                row_data = []
                for col in range(start_col, end_col + 1):
                    cell_value = sheet.cell(row, col).value
                    row_data.append(cell_value)
                data_matrix.append(row_data)
            
            # ヘッダー行を検出
            header_row_idx = self._detect_header_row(data_matrix)
            header_row = start_row + header_row_idx if header_row_idx is not None else None
            
            # ヘッダー取得
            headers = []
            if header_row_idx is not None:
                header_data = data_matrix[header_row_idx]
                headers = [str(cell) if cell is not None else f"列{i+1}" for i, cell in enumerate(header_data)]
            else:
                headers = [f"列{i+1}" for i in range(col_count)]
            
            # サンプルデータ取得（ヘッダー + 3行）
            sample_data = []
            data_start_idx = (header_row_idx + 1) if header_row_idx is not None else 0
            
            for i in range(data_start_idx, min(data_start_idx + 3, len(data_matrix))):
                if i < len(data_matrix):
                    row_dict = {}
                    for j, header in enumerate(headers):
                        if j < len(data_matrix[i]):
                            value = data_matrix[i][j]
                            row_dict[header] = str(value) if value is not None else ""
                        else:
                            row_dict[header] = ""
                    sample_data.append(row_dict)
            
            # データ密度計算
            total_cells = row_count * col_count
            data_cells = sum(1 for row in data_matrix for cell in row 
                           if cell is not None and str(cell).strip() != '')
            data_density = data_cells / total_cells if total_cells > 0 else 0
            
            # 品質スコア計算
            quality_score = self._calculate_quality_score(
                data_matrix, row_count, col_count, data_density, header_row_idx
            )
            
            # 推定レコード数
            estimated_records = row_count - 1 if header_row_idx is not None else row_count
            
            return TableCandidate(
                table_id=table_id,
                sheet_name=sheet_name,
                start_row=start_row,
                end_row=end_row,
                start_col=start_col,
                end_col=end_col,
                header_row=header_row,
                quality_score=quality_score,
                data_density=data_density,
                row_count=row_count,
                col_count=col_count,
                estimated_records=estimated_records,
                headers=headers,
                sample_data=sample_data,
                metadata={
                    'detection_method': 'statistical',
                    'data_cells': data_cells,
                    'total_cells': total_cells
                }
            )
            
        except Exception as e:
            logger.error(f"Error analyzing data region: {str(e)}")
            return None
    
    def _detect_header_row(self, data_matrix) -> Optional[int]:
        """ヘッダー行を検出する"""
        if len(data_matrix) < 2:
            return None
        
        best_header_idx = None
        best_score = 0
        
        # 最初の3行までをヘッダー候補として分析
        for row_idx in range(min(3, len(data_matrix))):
            row = data_matrix[row_idx]
            
            # 文字列の割合を計算
            string_count = 0
            non_empty_count = 0
            
            for cell in row:
                if cell is not None and str(cell).strip() != '':
                    non_empty_count += 1
                    # 数値でない場合は文字列とみなす
                    try:
                        float(str(cell).replace(',', ''))
                    except (ValueError, TypeError):
                        string_count += 1
            
            if non_empty_count == 0:
                continue
            
            string_ratio = string_count / non_empty_count
            
            # ヘッダースコア（文字列率 + データ充填率）
            fill_ratio = non_empty_count / len(row)
            header_score = string_ratio * 0.7 + fill_ratio * 0.3
            
            if header_score > best_score and string_ratio >= 0.5:
                best_score = header_score
                best_header_idx = row_idx
        
        return best_header_idx
    
    def _calculate_quality_score(self, data_matrix, row_count, col_count, data_density, header_row_idx) -> float:
        """表の品質スコアを計算"""
        score = 0.0
        
        # データ密度による評価 (0-0.3)
        density_score = min(data_density * 3, 0.3)
        score += density_score
        
        # サイズによる評価 (0-0.3)
        size_score = min((row_count * col_count) / 100, 0.3)
        score += size_score
        
        # ヘッダーの有無による評価 (0-0.2)
        if header_row_idx is not None:
            score += 0.2
        
        # データの一貫性による評価 (0-0.2)
        consistency_score = self._calculate_data_consistency(data_matrix, header_row_idx)
        score += consistency_score
        
        return min(score, 1.0)
    
    def _calculate_data_consistency(self, data_matrix, header_row_idx) -> float:
        """データの一貫性スコアを計算"""
        if len(data_matrix) <= 1:
            return 0.0
        
        data_start = (header_row_idx + 1) if header_row_idx is not None else 0
        if data_start >= len(data_matrix):
            return 0.0
        
        col_consistency_scores = []
        
        for col_idx in range(len(data_matrix[0])):
            col_data = []
            for row_idx in range(data_start, len(data_matrix)):
                if col_idx < len(data_matrix[row_idx]):
                    cell = data_matrix[row_idx][col_idx]
                    if cell is not None and str(cell).strip() != '':
                        col_data.append(cell)
            
            if len(col_data) < 2:
                continue
            
            # データ型の一貫性を確認
            numeric_count = 0
            date_count = 0
            
            for value in col_data:
                str_value = str(value).strip()
                
                # 数値チェック
                try:
                    float(str_value.replace(',', ''))
                    numeric_count += 1
                    continue
                except (ValueError, TypeError):
                    pass
                
                # 日付チェック
                if re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', str_value):
                    date_count += 1
            
            total_values = len(col_data)
            if total_values > 0:
                numeric_ratio = numeric_count / total_values
                date_ratio = date_count / total_values
                
                # 一貫性スコア（1種類のデータ型が80%以上なら高評価）
                if numeric_ratio >= 0.8 or date_ratio >= 0.8:
                    col_consistency_scores.append(1.0)
                elif numeric_ratio >= 0.6 or date_ratio >= 0.6:
                    col_consistency_scores.append(0.6)
                else:
                    col_consistency_scores.append(0.3)
        
        if not col_consistency_scores:
            return 0.0
        
        avg_consistency = sum(col_consistency_scores) / len(col_consistency_scores)
        return min(avg_consistency * 0.2, 0.2)

# グローバル表検出器インスタンス
default_table_detector = StatisticalTableDetector()

# セッション管理用のグローバル変数
session_data: Dict[str, Dict[str, Any]] = {}
session_timestamps: Dict[str, datetime] = {}

# セッション有効期限（30分）
SESSION_TIMEOUT = timedelta(minutes=30)

class SessionData:
    """セッションデータを管理するクラス"""
    def __init__(self):
        self.raw_data: Optional[pd.DataFrame] = None
        self.processed_data: Optional[pd.DataFrame] = None
        self.analysis_result: Dict[str, Any] = {}
        self.metadata: Dict[str, Any] = {}
        self.file_info: Dict[str, Any] = {}

def cleanup_expired_sessions():
    """期限切れのセッションをクリーンアップ"""
    current_time = datetime.now()
    expired_sessions = []
    
    for session_id, timestamp in session_timestamps.items():
        if current_time - timestamp > SESSION_TIMEOUT:
            expired_sessions.append(session_id)
    
    for session_id in expired_sessions:
        if session_id in session_data:
            del session_data[session_id]
        if session_id in session_timestamps:
            del session_timestamps[session_id]
        logger.info(f"Expired session cleaned up: {session_id}")
    
    return len(expired_sessions)

def get_session_data(session_id: str) -> Optional[Dict[str, Any]]:
    """セッションデータを取得"""
    cleanup_expired_sessions()
    
    if session_id not in session_data:
        return None
    
    # アクセス時刻を更新
    session_timestamps[session_id] = datetime.now()
    return session_data[session_id]

def create_session(session_id: str) -> Dict[str, Any]:
    """新しいセッションを作成"""
    cleanup_expired_sessions()
    
    session_data[session_id] = {
        'raw_data': None,
        'processed_data': None,
        'analysis_result': {},
        'metadata': {},
        'file_info': {}
    }
    session_timestamps[session_id] = datetime.now()
    return session_data[session_id]

# 許可されるファイルタイプ
ALLOWED_EXTENSIONS = {'.xlsx', '.xls', '.csv'}
ALLOWED_MIME_TYPES = {
    'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
    'application/vnd.ms-excel',  # .xls
    'text/csv',  # .csv
    'application/csv'  # .csv
}

# ファイルサイズ制限（10MB）
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB in bytes

def detect_file_type(file: UploadFile) -> str:
    """ファイルの形式を判定する"""
    if not file.filename:
        raise HTTPException(status_code=400, detail="ファイル名が指定されていません")
    
    # ファイル拡張子チェック
    file_extension = os.path.splitext(file.filename.lower())[1]
    if file_extension not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400, 
            detail=f"サポートされていないファイル形式です: {file_extension}"
        )
    
    # MIMEタイプチェック
    if file.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=400, 
            detail=f"サポートされていないMIMEタイプです: {file.content_type}"
        )
    
    # ファイルタイプを返す
    if file_extension == '.csv':
        return "csv"
    elif file_extension in ['.xlsx', '.xls']:
        return "excel"
    else:
        raise HTTPException(status_code=400, detail="未知のファイル形式です")

def validate_file_size(content: bytes) -> None:
    """ファイルサイズを検証する"""
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"ファイルサイズが大きすぎます（最大10MB）: {len(content)} bytes"
        )

def detect_header_row(df: pd.DataFrame) -> int:
    """CSVファイルのヘッダー行を検出する"""
    try:
        # 最初の10行を分析（ファイルが小さい場合は全行）
        max_rows_to_check = min(10, len(df))
        
        header_candidates = []
        
        for i in range(max_rows_to_check):
            row = df.iloc[i]
            
            # 空行をスキップ
            if row.isna().all() or (row == '').all():
                continue
            
            # 文字列データの割合を計算
            string_count = 0
            numeric_count = 0
            
            for value in row:
                if pd.isna(value) or value == '':
                    continue
                
                # 数値かどうかをチェック
                try:
                    float(str(value).replace(',', ''))
                    numeric_count += 1
                except (ValueError, TypeError):
                    string_count += 1
            
            total_values = string_count + numeric_count
            if total_values == 0:
                continue
            
            string_ratio = string_count / total_values
            
            # ヘッダー行の候補として評価
            # - 文字列の割合が高い（70%以上）
            # - 少なくとも2つ以上の値がある
            if string_ratio >= 0.7 and total_values >= 2:
                header_candidates.append({
                    'row_index': i,
                    'string_ratio': string_ratio,
                    'total_values': total_values
                })
        
        # 最も文字列の割合が高い行をヘッダーとして選択
        if header_candidates:
            best_candidate = max(header_candidates, key=lambda x: x['string_ratio'])
            return best_candidate['row_index']
        
        # ヘッダー候補が見つからない場合は0行目を返す
        return 0
        
    except Exception as e:
        logger.warning(f"Header detection failed: {e}, using row 0 as default")
        return 0

def analyze_data_types(df: pd.DataFrame) -> Dict[str, str]:
    """データ型を詳細に分析する"""
    data_types = {}
    
    for col in df.columns:
        col_data = df[col].dropna()  # 欠損値を除外
        
        if len(col_data) == 0:
            data_types[col] = 'empty'
            continue
        
        # 数値型の判定
        numeric_count = 0
        date_count = 0
        boolean_count = 0
        
        for value in col_data:
            str_value = str(value).strip()
            
            # Boolean型チェック
            if str_value.lower() in ['true', 'false', 'yes', 'no', 'y', 'n', '1', '0']:
                boolean_count += 1
                continue
            
            # 数値型チェック（カンマ区切りの数値も考慮）
            try:
                float(str_value.replace(',', ''))
                numeric_count += 1
                continue
            except (ValueError, TypeError):
                pass
            
            # 日付型チェック
            date_patterns = [
                r'\d{4}-\d{2}-\d{2}',  # YYYY-MM-DD
                r'\d{4}/\d{2}/\d{2}',  # YYYY/MM/DD
                r'\d{2}/\d{2}/\d{4}',  # MM/DD/YYYY
                r'\d{2}-\d{2}-\d{4}',  # MM-DD-YYYY
            ]
            
            for pattern in date_patterns:
                if re.match(pattern, str_value):
                    date_count += 1
                    break
        
        total_values = len(col_data)
        
        # データ型を決定（閾値: 80%）
        if boolean_count / total_values >= 0.8:
            data_types[col] = 'boolean'
        elif date_count / total_values >= 0.8:
            data_types[col] = 'date'
        elif numeric_count / total_values >= 0.8:
            # 整数か小数点数かを判定
            integer_count = 0
            for value in col_data:
                str_value = str(value).strip().replace(',', '')
                try:
                    float_val = float(str_value)
                    if float_val.is_integer():
                        integer_count += 1
                except (ValueError, TypeError):
                    pass
            
            if integer_count / numeric_count >= 0.9:
                data_types[col] = 'integer'
            else:
                data_types[col] = 'number'
        else:
            # カテゴリ型かどうかを判定
            unique_ratio = len(col_data.unique()) / total_values
            if unique_ratio <= 0.1 and len(col_data.unique()) <= 20:  # 重複が多くユニーク値が少ない
                data_types[col] = 'category'
            else:
                data_types[col] = 'string'
    
    return data_types

def get_excel_sheets_info(file_content: bytes) -> List[Dict[str, Any]]:
    """Excelファイルからシート情報を取得"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(file_content), read_only=True)
        sheets_info = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # シートの最大行・列を取得
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # データの有無を確認（最初の100行をサンプリング）
            has_data = False
            data_cells = 0
            sample_rows = min(100, max_row) if max_row else 0
            sample_cols = min(20, max_col) if max_col else 0
            
            if sample_rows > 0 and sample_cols > 0:
                for row in range(1, sample_rows + 1):
                    for col in range(1, sample_cols + 1):
                        cell_value = sheet.cell(row, col).value
                        if cell_value is not None and str(cell_value).strip() != '':
                            data_cells += 1
                            has_data = True
                
                # データ密度を計算
                total_sample_cells = sample_rows * sample_cols
                data_density = data_cells / total_sample_cells if total_sample_cells > 0 else 0
            else:
                data_density = 0
            
            # データ範囲を推定
            data_range = None
            if has_data and max_row > 0 and max_col > 0:
                end_col_letter = openpyxl.utils.get_column_letter(max_col)
                data_range = f"A1:{end_col_letter}{max_row}"
            
            sheet_info = {
                "name": sheet_name,
                "row_count": max_row if max_row else 0,
                "col_count": max_col if max_col else 0,
                "has_data": has_data,
                "data_range": data_range,
                "data_density": round(data_density, 3),
                "estimated_data_cells": data_cells
            }
            sheets_info.append(sheet_info)
        
        workbook.close()
        return sheets_info
        
    except Exception as e:
        logger.error(f"Error reading Excel sheets: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Excelファイルのシート情報取得中にエラーが発生しました: {str(e)}"
        )

def validate_excel_file(file: UploadFile) -> bool:
    """Excelファイルの妥当性をチェック"""
    if not file.filename:
        return False
    
    # ファイル拡張子チェック
    file_extension = os.path.splitext(file.filename.lower())[1]
    if file_extension not in ['.xlsx', '.xls']:
        return False
    
    # MIMEタイプチェック
    excel_mime_types = {
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  # .xlsx
        'application/vnd.ms-excel',  # .xls
    }
    if file.content_type not in excel_mime_types:
        return False
    
    return True

def process_excel_sheets(file_content: bytes, filename: str, session_id: str) -> Dict[str, Any]:
    """Excelファイルのシート情報を処理してセッションに保存"""
    try:
        logger.info(f"Processing Excel sheets for file: {filename} (session: {session_id})")
        
        # セッションデータを作成
        session = create_session(session_id)
        
        # シート情報を取得
        sheets_info = get_excel_sheets_info(file_content)
        
        # ワークブックを保存（後でシート選択時に使用）
        workbook = openpyxl.load_workbook(BytesIO(file_content))
        
        # セッションにデータを保存
        session['file_info'] = {
            'filename': filename,
            'file_type': 'excel',
            'total_sheets': len(sheets_info),
            'sheets_info': sheets_info
        }
        
        # ワークブックデータも保存（バイト形式で保存）
        session['raw_workbook_data'] = file_content
        
        return {
            "filename": filename,
            "sheets": sheets_info,
            "total_sheets": len(sheets_info)
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file {filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Excelファイルの処理中にエラーが発生しました: {str(e)}"
        )

def analyze_data_quality(df: pd.DataFrame) -> Dict[str, Any]:
    """データ品質を分析する"""
    quality_report = {
        'total_rows': len(df),
        'total_columns': len(df.columns),
        'missing_data': {},
        'data_consistency': {},
        'statistics': {}
    }
    
    for col in df.columns:
        col_data = df[col]
        
        # 欠損値分析
        missing_count = col_data.isna().sum()
        missing_ratio = missing_count / len(col_data) if len(col_data) > 0 else 0
        
        quality_report['missing_data'][col] = {
            'count': int(missing_count),
            'ratio': float(missing_ratio)
        }
        
        # データ一貫性分析
        non_missing_data = col_data.dropna()
        if len(non_missing_data) > 0:
            unique_count = len(non_missing_data.unique())
            unique_ratio = unique_count / len(non_missing_data)
            
            quality_report['data_consistency'][col] = {
                'unique_values': int(unique_count),
                'unique_ratio': float(unique_ratio),
                'duplicates': int(len(non_missing_data) - unique_count)
            }
            
            # 数値列の統計情報
            try:
                numeric_data = pd.to_numeric(non_missing_data, errors='coerce').dropna()
                if len(numeric_data) > 0:
                    quality_report['statistics'][col] = {
                        'mean': float(numeric_data.mean()),
                        'median': float(numeric_data.median()),
                        'std': float(numeric_data.std()),
                        'min': float(numeric_data.min()),
                        'max': float(numeric_data.max())
                    }
            except Exception:
                pass  # 数値変換できない場合はスキップ
    
    return quality_report

async def process_csv_advanced(file_content: bytes, filename: str, session_id: str) -> Dict[str, Any]:
    """CSV ファイルを高度に解析する"""
    try:
        logger.info(f"Advanced processing CSV file: {filename} (session: {session_id})")
        
        # セッションデータを作成
        session = create_session(session_id)
        
        # CSVファイルを読み込み（ヘッダーなしで読み込んで後で判定）
        df_raw = pd.read_csv(BytesIO(file_content), header=None)
        
        # ヘッダー行を検出
        header_row = detect_header_row(df_raw)
        
        # ヘッダー行を使用してデータフレームを再構成
        if header_row > 0:
            # ヘッダー行より前の行を削除
            df_raw = df_raw.iloc[header_row:].reset_index(drop=True)
        
        # 1行目を列名として使用
        df = pd.read_csv(BytesIO(file_content), skiprows=header_row)
        
        # セッションにデータを保存
        session['raw_data'] = df_raw
        session['processed_data'] = df
        session['file_info'] = {
            'filename': filename,
            'file_type': 'csv',
            'detected_header_row': header_row,
            'total_rows': len(df),
            'total_columns': len(df.columns)
        }
        
        # 詳細なデータ型分析
        data_types = analyze_data_types(df)
        
        # データ品質分析
        quality_report = analyze_data_quality(df)
        
        # 分析結果をセッションに保存
        session['analysis_result'] = {
            'data_types': data_types,
            'quality_report': quality_report
        }
        
        # 基本情報を取得
        columns = df.columns.tolist()
        total_rows = len(df)
        
        # サンプルデータを取得（最初の5行）
        sample_data = df.head(5).fillna('').to_dict('records')
        
        return {
            "file_type": "csv",
            "detected_header_row": header_row,
            "columns": columns,
            "data_types": data_types,
            "sample_data": sample_data,
            "total_rows": total_rows,
            "quality_summary": {
                "missing_data_columns": len([col for col, info in quality_report['missing_data'].items() if info['ratio'] > 0]),
                "numeric_columns": len([col for col, dtype in data_types.items() if dtype in ['integer', 'number']]),
                "date_columns": len([col for col, dtype in data_types.items() if dtype == 'date']),
                "category_columns": len([col for col, dtype in data_types.items() if dtype == 'category'])
            }
        }
        
    except Exception as e:
        logger.error(f"Error processing CSV file {filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"CSVファイルの処理中にエラーが発生しました: {str(e)}"
        )

# 後方互換性のため古い関数も残しておく
async def process_csv(file_content: bytes, filename: str) -> Dict[str, Any]:
    """CSV ファイルを処理する（シンプル版）"""
    session_id = str(uuid.uuid4())
    result = await process_csv_advanced(file_content, filename, session_id)
    return result

async def process_excel(file_content: bytes, filename: str) -> Dict[str, Any]:
    """Excel ファイルを処理する（非正規化ファイルを想定）"""
    try:
        logger.info(f"Processing Excel file: {filename}")
        
        # 現段階では基本的な読み込みのみ実装
        # 後でテーブル走査機能を追加予定
        df = pd.read_excel(BytesIO(file_content))
        
        # 基本情報を取得
        columns = df.columns.tolist()
        total_rows = len(df)
        
        # データ型を推論
        data_types = {}
        for col in columns:
            dtype = str(df[col].dtype)
            # pandasのデータ型を簡潔な形式に変換
            if dtype.startswith('int'):
                data_types[col] = 'integer'
            elif dtype.startswith('float'):
                data_types[col] = 'number'
            elif dtype.startswith('bool'):
                data_types[col] = 'boolean'
            elif dtype.startswith('datetime'):
                data_types[col] = 'date'
            else:
                data_types[col] = 'string'
        
        # サンプルデータを取得（最初の5行）
        sample_data = df.head(5).to_dict('records')
        
        return {
            "file_type": "excel",
            "detected_header_row": 0,  # 暫定的に0行目を設定（後で改善予定）
            "columns": columns,
            "data_types": data_types,
            "sample_data": sample_data,
            "total_rows": total_rows,
            "note": "Excel表走査機能は次段階で実装予定"
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file {filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Excelファイルの処理中にエラーが発生しました: {str(e)}"
        )

@router.post("/parse-excel")
async def parse_excel(file: UploadFile = File(...)):
    """Excel/CSVファイルを解析するエンドポイント"""
    try:
        # ファイル内容を読み取り
        file_content = await file.read()
        
        # ファイルサイズを検証
        validate_file_size(file_content)
        
        # ファイル形式を判定
        file_type = detect_file_type(file)
        
        # セッションIDを生成
        session_id = str(uuid.uuid4())
        
        logger.info(f"Processing file: {file.filename}, type: {file_type}, session: {session_id}")
        
        # ファイル形式に応じて処理を分岐
        if file_type == "csv":
            result = await process_csv_advanced(file_content, file.filename, session_id)
        elif file_type == "excel":
            result = await process_excel(file_content, file.filename)
        else:
            raise HTTPException(status_code=400, detail="サポートされていないファイル形式です")
        
        # セッションIDを結果に追加
        result["session_id"] = session_id
        result["filename"] = file.filename
        
        return {
            "status": "success",
            "message": "ファイルの解析が完了しました",
            "data": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error processing file {file.filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"ファイル処理中に予期しないエラーが発生しました: {str(e)}"
        )

@router.get("/session/{session_id}/data")
async def get_session_data_detail(session_id: str = Path(...)):
    """セッションの詳細データを取得"""
    session = get_session_data(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="セッションが見つからないか期限切れです")
    
    try:
        # DataFrameをJSON形式に変換
        processed_data = session.get('processed_data')
        if processed_data is not None:
            # 全データを返す（大きなファイルの場合は制限をかけることも可能）
            data_dict = processed_data.fillna('').to_dict('records')
            return {
                "status": "success",
                "session_id": session_id,
                "file_info": session.get('file_info', {}),
                "data": data_dict,
                "total_rows": len(data_dict)
            }
        else:
            raise HTTPException(status_code=404, detail="処理済みデータが見つかりません")
    
    except Exception as e:
        logger.error(f"Error retrieving session data {session_id}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"セッションデータの取得中にエラーが発生しました: {str(e)}"
        )

@router.get("/session/{session_id}/analysis")
async def get_session_analysis(session_id: str = Path(...)):
    """セッションの分析結果を取得"""
    session = get_session_data(session_id)
    if not session:
        raise HTTPException(status_code=404, detail="セッションが見つからないか期限切れです")
    
    try:
        analysis_result = session.get('analysis_result', {})
        file_info = session.get('file_info', {})
        
        return {
            "status": "success",
            "session_id": session_id,
            "file_info": file_info,
            "analysis": analysis_result
        }
    
    except Exception as e:
        logger.error(f"Error retrieving session analysis {session_id}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"分析結果の取得中にエラーが発生しました: {str(e)}"
        )

@router.delete("/session/{session_id}")
async def delete_session(session_id: str = Path(...)):
    """セッションを削除"""
    if session_id in session_data:
        del session_data[session_id]
    if session_id in session_timestamps:
        del session_timestamps[session_id]
    
    logger.info(f"Session deleted: {session_id}")
    return {
        "status": "success",
        "message": f"セッション {session_id} が削除されました"
    }

@router.post("/parse-excel-sheets")
async def parse_excel_sheets(file: UploadFile = File(...)):
    """Excelファイルのシート一覧を取得するエンドポイント"""
    try:
        # Excelファイルの妥当性をチェック
        if not validate_excel_file(file):
            raise HTTPException(
                status_code=400,
                detail="サポートされていないファイル形式です。Excel形式（.xlsx, .xls）のファイルをアップロードしてください。"
            )
        
        # ファイル内容を読み取り
        file_content = await file.read()
        
        # ファイルサイズを検証
        validate_file_size(file_content)
        
        # セッションIDを生成
        session_id = str(uuid.uuid4())
        
        logger.info(f"Processing Excel sheets for: {file.filename}, session: {session_id}")
        
        # Excelシート情報を処理
        result = process_excel_sheets(file_content, file.filename, session_id)
        
        return {
            "status": "success",
            "message": "Excelファイルのシート一覧を取得しました",
            "session_id": session_id,
            "data": result
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Unexpected error processing Excel file {file.filename}: {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"Excelファイル処理中に予期しないエラーが発生しました: {str(e)}"
        )

@router.get("/sessions")
async def list_active_sessions():
    """アクティブなセッション一覧を取得"""
    cleanup_expired_sessions()
    
    sessions_info = []
    for session_id, timestamp in session_timestamps.items():
        session = session_data.get(session_id, {})
        file_info = session.get('file_info', {})
        
        sessions_info.append({
            "session_id": session_id,
            "created_at": timestamp.isoformat(),
            "filename": file_info.get('filename', 'Unknown'),
            "file_type": file_info.get('file_type', 'Unknown'),
            "total_rows": file_info.get('total_rows', 0),
            "total_columns": file_info.get('total_columns', 0),
            "total_sheets": file_info.get('total_sheets', 0)
        })
    
    return {
        "status": "success",
        "active_sessions": sessions_info,
        "total_count": len(sessions_info)
    }

@router.post("/excel-sheet-tables/{session_id}/{sheet_name}")
async def detect_tables_in_sheet(
    session_id: str = Path(...),
    sheet_name: str = Path(...),
    min_rows: int = 3,
    min_cols: int = 2,
    max_tables: int = 10
):
    """指定されたシート内の表を検出するエンドポイント"""
    try:
        # セッションを取得
        session = get_session_data(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="セッションが見つからないか期限切れです")
        
        # ワークブックデータを取得
        workbook_data = session.get('raw_workbook_data')
        if not workbook_data:
            raise HTTPException(status_code=404, detail="Excelワークブックデータが見つかりません")
        
        # シート名をデコード（URLエンコードされている可能性があるため）
        import urllib.parse
        decoded_sheet_name = urllib.parse.unquote(sheet_name)
        
        logger.info(f"Detecting tables in sheet '{decoded_sheet_name}' (session: {session_id})")
        
        # 表検出を実行
        table_candidates = default_table_detector.detect_tables(
            workbook_data=workbook_data,
            sheet_name=decoded_sheet_name,
            min_rows=min_rows,
            min_cols=min_cols,
            max_tables=max_tables
        )
        
        # 検出された表をセッションに保存
        session['detected_tables'] = {
            'sheet_name': decoded_sheet_name,
            'tables': [table.to_dict() for table in table_candidates],
            'detection_info': default_table_detector.get_detector_info(),
            'detection_time': datetime.now().isoformat()
        }
        
        # レスポンスデータを構築
        response_data = {
            "sheet_name": decoded_sheet_name,
            "total_tables": len(table_candidates),
            "tables": [table.to_dict() for table in table_candidates],
            "detection_info": default_table_detector.get_detector_info()
        }
        
        return {
            "status": "success",
            "message": f"シート '{decoded_sheet_name}' で {len(table_candidates)} 個の表を検出しました",
            "session_id": session_id,
            "data": response_data
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error detecting tables in sheet {sheet_name} (session: {session_id}): {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"表検出中にエラーが発生しました: {str(e)}"
        )

@router.post("/select-table/{session_id}/{table_id}")
async def select_table(
    session_id: str = Path(...),
    table_id: str = Path(...)
):
    """選択された表のデータを取得し、最終処理を行うエンドポイント"""
    try:
        # セッションを取得
        session = get_session_data(session_id)
        if not session:
            raise HTTPException(status_code=404, detail="セッションが見つからないか期限切れです")
        
        # 検出済み表データを取得
        detected_tables = session.get('detected_tables')
        if not detected_tables:
            raise HTTPException(status_code=404, detail="検出済み表データが見つかりません")
        
        # 指定された表を探す
        selected_table = None
        for table_data in detected_tables['tables']:
            if table_data['table_id'] == table_id:
                selected_table = table_data
                break
        
        if not selected_table:
            raise HTTPException(status_code=404, detail=f"表ID '{table_id}' が見つかりません")
        
        # ワークブックデータから指定範囲のデータを抽出
        workbook_data = session.get('raw_workbook_data')
        if not workbook_data:
            raise HTTPException(status_code=404, detail="Excelワークブックデータが見つかりません")
        
        # 表の全データを取得
        full_table_data = extract_table_data(
            workbook_data, 
            detected_tables['sheet_name'], 
            selected_table
        )
        
        # セッションに最終データを保存
        session['selected_table'] = {
            'table_info': selected_table,
            'full_data': full_table_data,
            'selection_time': datetime.now().isoformat()
        }
        
        # ファイル情報を更新
        file_info = session.get('file_info', {})
        file_info.update({
            'selected_sheet': detected_tables['sheet_name'],
            'selected_table_id': table_id,
            'final_rows': len(full_table_data['records']),
            'final_columns': len(full_table_data['headers'])
        })
        session['file_info'] = file_info
        
        return {
            "status": "success",
            "message": f"表 '{table_id}' を選択しました",
            "session_id": session_id,
            "data": {
                "table_info": selected_table,
                "headers": full_table_data['headers'],
                "total_records": len(full_table_data['records']),
                "sample_records": full_table_data['records'][:10],  # 最初の10件のみ
                "data_types": full_table_data.get('data_types', {}),
                "quality_info": full_table_data.get('quality_info', {})
            }
        }
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error selecting table {table_id} (session: {session_id}): {str(e)}")
        raise HTTPException(
            status_code=500,
            detail=f"表選択中にエラーが発生しました: {str(e)}"
        )

def extract_table_data(workbook_data: bytes, sheet_name: str, table_info: Dict[str, Any]) -> Dict[str, Any]:
    """指定された表の全データを抽出する"""
    try:
        workbook = openpyxl.load_workbook(BytesIO(workbook_data))
        sheet = workbook[sheet_name]
        
        range_info = table_info['range']
        start_row = range_info['start_row']
        end_row = range_info['end_row']
        start_col = range_info['start_col']
        end_col = range_info['end_col']
        header_row = table_info.get('header_row')
        headers = table_info.get('headers', [])
        
        # 全データを収集
        all_data = []
        for row in range(start_row, end_row + 1):
            row_data = []
            for col in range(start_col, end_col + 1):
                cell_value = sheet.cell(row, col).value
                row_data.append(cell_value)
            all_data.append(row_data)
        
        # ヘッダー行がある場合はデータ部分のみ抽出
        records = []
        data_start_idx = 1 if header_row else 0
        
        for row_idx in range(data_start_idx, len(all_data)):
            record = {}
            for col_idx, header in enumerate(headers):
                if col_idx < len(all_data[row_idx]):
                    value = all_data[row_idx][col_idx]
                    record[header] = str(value) if value is not None else ""
                else:
                    record[header] = ""
            records.append(record)
        
        # データ型分析
        if records:
            df = pd.DataFrame(records)
            data_types = analyze_data_types(df)
            quality_report = analyze_data_quality(df)
        else:
            data_types = {}
            quality_report = {}
        
        workbook.close()
        
        return {
            'headers': headers,
            'records': records,
            'data_types': data_types,
            'quality_info': quality_report,
            'total_records': len(records)
        }
        
    except Exception as e:
        logger.error(f"Error extracting table data: {str(e)}")
        raise